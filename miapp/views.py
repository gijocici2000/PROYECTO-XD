import datetime
import io
import json
from decimal import Decimal
import os
from typing import Optional
from django.db import transaction
from django.db.models import Q, Sum
from django.forms import formset_factory, inlineformset_factory
from django.http import HttpRequest, HttpResponse, FileResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.serializers.json import DjangoJSONEncoder
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.http import HttpResponse
from django.urls import reverse
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT,TA_LEFT
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph
from reportlab.pdfgen import canvas
from django.utils.safestring import mark_safe
from .models import *
from .forms import *
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils.timezone import now
from datetime import datetime
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.utils import simpleSplit


#################################login y registro de usuarios##########################################


@login_required
def index(request):
    cargos_permitidos = ['admin', 'gerente']

    try:
        cargo_usuario = request.user.empleado.cargo.nombre.lower()
    except Empleado.DoesNotExist:
        cargo_usuario = ''

    total_productos = Producto.objects.count()
    total_clientes = Cliente.objects.count()
    total_empleados = Empleado.objects.count()

    return render(request, 'index.html', {
        'cargos_permitidos': cargos_permitidos,
        'cargo_usuario': cargo_usuario,
        'total_productos': total_productos,
        'total_clientes': total_clientes,
        'total_empleados': total_empleados
    })
def cargo_requerido(nombre_cargos):
    def decorador(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')

            try:
                empleado = Empleado.objects.get(user=request.user)
                if empleado.cargo.nombre.lower() in [c.lower() for c in nombre_cargos]:
                    return view_func(request, *args, **kwargs)
            except Empleado.DoesNotExist:
                pass

            return redirect('no_autorizado')
        return wrapper
    return decorador


def register_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()  # Aquí crea User y Empleado juntos
            messages.success(request, "Usuario registrado correctamente.")
            return redirect('login')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registro_usuario/register.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('index')  
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    else:
        form = AuthenticationForm()
    return render(request, 'registro_usuario/login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('index')
def no_autorizado(request):
    return render(request, 'registro_usuario/no_autorizado.html', {"mensaje": "No tienes permiso para acceder aquí."})

######################################### EMPLEADO ########################################

@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def crear_empleado(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = EmpleadoForm(request.POST, request.FILES)
        if form.is_valid():
            empleado = form.save(commit=False)
            empleado.creacion_usuario = request.user.username
            empleado.save()
            return redirect("consultar_empleado")
    else:
        form = EmpleadoForm()
    return render(request, "empleado/crear_empleado.html", {"form": form, "edit_mode": False})
 
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])   
def modificar_empleado(request: HttpRequest, id: int) -> HttpResponse:
    empleado = get_object_or_404(Empleado, id=id)
    if request.method == "POST":
        form = EmpleadoForm(request.POST, request.FILES, instance=empleado)
        if form.is_valid():
            empleado = form.save(commit=False)
            empleado.modificacion_usuario = request.user.username
            empleado.save()
            return redirect("consultar_empleado")
    else:
        form = EmpleadoForm(instance=empleado)
    return render(request, "empleado/modificar_empleado.html", {"form": form, "empleado": empleado, "edit_mode": True})
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def consultar_empleado(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        if 'buscar_empleado' in request.POST:
            buscarEmpleadoForm = BuscarEmpleadoForm(request.POST)
            if buscarEmpleadoForm.is_valid():
                nombre = buscarEmpleadoForm.cleaned_data.get('nombre', '')
                apellido = buscarEmpleadoForm.cleaned_data.get('apellido', '')
                cedula = buscarEmpleadoForm.cleaned_data.get('cedula', '')
                fecha = buscarEmpleadoForm.cleaned_data.get('fecha', None)

                empleados = Empleado.objects.filter(
                    Q(nombre__icontains=nombre) if nombre else Q(),
                    Q(apellido__icontains=apellido) if apellido else Q(),
                    Q(cedula__icontains=cedula) if cedula else Q(),
                    Q(fecha_creacion__lte=fecha) if fecha else Q(),
                    
                )
                context = {'buscador_Empleado': buscarEmpleadoForm, 'empleados': empleados}
                return render(request, "empleado/consultar_empleado.html", context)

    buscarEmpleadoForm = BuscarEmpleadoForm()
    empleados = None
    return render(request, "empleado/consultar_empleado.html", {'buscador_Empleado': buscarEmpleadoForm, 'empleados': empleados})
@login_required
@cargo_requerido(['admin', 'Gerente', 'supervisor'])
def eliminar_empleado(request: HttpRequest, id: int) -> HttpResponse:
    empleado = get_object_or_404(Empleado, id=id)
    if request.method == "POST":
        empleado.estado = 0  # Marcar como inactivo
        empleado.save()
        return redirect("consultar_empleado")
    return render(request, "empleado/eliminar_empleado.html", {"empleado": empleado})




########################################## CLIENTE ########################################
@login_required
@cargo_requerido(['admin', 'gerente','cajero','supervisor' ,])
def crear_cliente(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = ClienteForm(data=request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            return redirect("consultar_cliente")
    else:
        form = ClienteForm()
    context = {"form_Cliente": form, "edit_mode": False}
    return render(request, "facturacion_cliente/cliente/crear_cliente.html", context)

@login_required
@cargo_requerido(['admin', 'gerente', 'cajero', 'supervisor'])
def consultar_cliente(request: HttpRequest) -> HttpResponse:

    buscarClienteForm = BuscarClienteForm(request.POST or None)

    # Mostrar todos los clientes por defecto
    clientes = Cliente.objects.all()

    if request.method == "POST":
        if buscarClienteForm.is_valid():
            nombre = buscarClienteForm.cleaned_data.get('nombre', '').strip()
            apellido = buscarClienteForm.cleaned_data.get('apellido', '').strip()
            cedula = buscarClienteForm.cleaned_data.get('cedula', '').strip()
            estado = buscarClienteForm.cleaned_data.get('estado', '').strip()

            filtros = Q()

            if nombre:
                filtros &= Q(nombre__icontains=nombre)
            if apellido:
                filtros &= Q(apellido__icontains=apellido)
            if cedula:
                filtros &= Q(cedula__icontains=cedula)
            if estado != '':
                try:
                    estado_int = int(estado)
                    filtros &= Q(estado=estado_int)
                except ValueError:
                    pass  # Si por alguna razón no es un número válido, ignora el filtro

            clientes = clientes.filter(filtros)

    context = {
        'buscador_Cliente': buscarClienteForm,
        'clientes': clientes
    }
    return render(request, "facturacion_cliente/cliente/consultar_cliente.html", context)   
@login_required
@cargo_requerido(['Admin', 'Gerente','cajero','supervisor'])
def modificar_cliente(request: HttpRequest, id: int) -> HttpResponse:
    cliente = get_object_or_404(Cliente, id=id)
    if request.method == "POST":
        form = ClienteForm(data=request.POST, files=request.FILES, instance=cliente)
        if form.is_valid():
            form.save()
            return redirect("consultar_cliente")
    else:
        form = ClienteForm(instance=cliente)
    context = {"form": form, "cliente": cliente, "edit_mode": True}
    return render(request, "facturacion_cliente/cliente/modificar_cliente.html", context)

def eliminar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)

    if request.method == 'POST':
        cliente.estado = 0  # Marcar como inactivo
        cliente.save()
        return redirect('consultar_cliente')

    return render(request, 'facturacion_cliente/cliente/eliminar_cliente.html', {'cliente': cliente})

############################################ PRODUCTO ########################################

@login_required
@cargo_requerido(['Admin', 'Gerente'])
def crear_producto(request):
    mensaje = ""
    stock_actual = None
    mostrar_stock = False  # por defecto oculto

    # Manejo AJAX para autocompletar por número de serie
    if request.method == 'GET' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        numero_serie = request.GET.get('numero_serie', '').strip()
        try:
            producto = Producto.objects.get(numero_serie__iexact=numero_serie)
            data = {
                'existe': True,
                'nombre': producto.nombre,
                'precio': float(producto.precio),
                'stock': producto.stock,
                'modelo': producto.modelo,
                'color': producto.color,
                'categoria': producto.categoria,
                'lote': producto.lote,
                'bodega': producto.bodega.nombre if producto.bodega else None,
                'imagen': producto.imagen.url if producto.imagen else '',
            }
        except Producto.DoesNotExist:
            data = {'existe': False}
        return JsonResponse(data)

    # Manejo POST (crear o actualizar producto)
    if request.method == 'POST':
        numero_serie = request.POST.get('numero_serie', '').strip()
        try:
            cantidad = int(request.POST.get('cantidad_ingresar', '0'))
        except (ValueError, TypeError):
            cantidad = 0

        if cantidad <= 0:
            mensaje = "❌ Ingresa una cantidad válida mayor a cero."
            form = ProductoForm(request.POST, request.FILES)  # aún si es inválido, pasar datos al form
        else:
            try:
                # Producto existente
                producto = Producto.objects.get(numero_serie__iexact=numero_serie)
                form = ProductoForm(request.POST, request.FILES, instance=producto)
                if form.is_valid():
                    prod = form.save(commit=False)
                    prod.stock = (prod.stock or 0) + cantidad
                    prod.modificacion_usuario = request.user.username
                    prod.save()
                    mensaje = f"✅ Producto existente actualizado. Nuevo stock: {prod.stock}"
                    stock_actual = prod.stock
                    mostrar_stock = True
                else:
                    mensaje = "❌ Formulario inválido. Revisa los datos."
                    print(form.errors)
            except Producto.DoesNotExist:
                # Producto nuevo
                form = ProductoForm(request.POST, request.FILES)
                if form.is_valid():
                    nuevo = form.save(commit=False)
                    nuevo.stock = cantidad
                    nuevo.creacion_usuario = request.user.username
                    nuevo.modificacion_usuario = request.user.username
                    nuevo.estado = 1
                    nuevo.save()
                    mensaje = "✅ Producto nuevo creado correctamente."
                    stock_actual = nuevo.stock
                    mostrar_stock = True
                    form = ProductoForm()  # limpiar formulario
                else:
                    mensaje = "❌ Formulario inválido. Revisa los datos."
                    print(form.errors)
    else:
        # Método GET sin AJAX: precargar si existe número de serie
        form = ProductoForm()
        numero_serie = request.GET.get('numero_serie', '').strip()
        if numero_serie:
            try:
                producto = Producto.objects.get(numero_serie__iexact=numero_serie)
                stock_actual = producto.stock
                mostrar_stock = True
                form = ProductoForm(instance=producto)
            except Producto.DoesNotExist:
                pass  # Producto nuevo, formulario vacío

    return render(request, 'facturacion_cliente/producto/crear_producto.html', {
        'form_Producto': form,
        'mensaje': mensaje,
        'stock_actual': stock_actual,
        'mostrar_stock': mostrar_stock,
    })

@login_required
@cargo_requerido(['Admin', 'Gerente', 'cajero', 'supervisor'])
def consultar_producto(request):

    productos = Producto.objects.all()
    if request.method == 'POST':
            filtro = request.POST.get('filtro', '').strip()
            if filtro:
                productos = productos.filter(
                    Q(nombre__icontains=filtro) |
                    Q(numero_serie__icontains=filtro)
                )


    form = BuscarProductoForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        nombre = form.cleaned_data.get('nombre')
        categoria = form.cleaned_data.get('categoria')
        numero_serie = form.cleaned_data.get('numero_serie')
        fecha_creacion = form.cleaned_data.get('fecha_creacion')

        if nombre:
            productos = productos.filter(nombre__icontains=nombre)
        if categoria:
            productos = productos.filter(categoria__icontains=categoria)
        if numero_serie:
            productos = productos.filter(numero_serie__icontains=numero_serie)
        if fecha_creacion:
            productos = productos.filter(fecha_creacion=fecha_creacion)

    context = {
        'productos': productos,
        'buscador_Producto': form,
    }
    return render(request, 'facturacion_cliente/producto/consultar_producto.html', context)

@login_required
@cargo_requerido(['Admin', 'Gerente', 'cajero', 'supervisor'])
def exportar_excel_producto(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Estilos
    font_encabezado = Font(bold=True, color="FFFFFF")
    fill_encabezado = PatternFill("solid", fgColor="4F81BD")  # azul profesional
    alineacion_centrada = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    columnas = [
        'Nombre', 'Modelo', 'Número de Serie', 'Categoría', 'Color', 
        'Lote', 'Precio', 'Stock', 'Bodega', 'Fecha de Creación', 'Imagen'
    ]
    ws.append(columnas)

    # Aplicar estilo encabezado
    for col_num, columna in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_num)
        celda.font = font_encabezado
        celda.fill = fill_encabezado
        celda.alignment = alineacion_centrada
        celda.border = borde

    productos = Producto.objects.all()

    for idx, p in enumerate(productos, start=2):
        ws.cell(row=idx, column=1, value=p.nombre).alignment = alineacion_centrada
        ws.cell(row=idx, column=2, value=p.modelo).alignment = alineacion_centrada
        ws.cell(row=idx, column=3, value=p.numero_serie).alignment = alineacion_centrada
        ws.cell(row=idx, column=4, value=p.categoria).alignment = alineacion_centrada
        ws.cell(row=idx, column=5, value=p.color).alignment = alineacion_centrada
        ws.cell(row=idx, column=6, value=p.lote).alignment = alineacion_centrada
        ws.cell(row=idx, column=7, value=float(p.precio)).alignment = alineacion_centrada
        ws.cell(row=idx, column=8, value=p.stock).alignment = alineacion_centrada
        ws.cell(row=idx, column=9, value=p.bodega.nombre if p.bodega else '').alignment = alineacion_centrada
        ws.cell(row=idx, column=10, value=p.fecha_creacion.strftime('%Y-%m-%d') if p.fecha_creacion else '').alignment = alineacion_centrada

        # Aplicar bordes a datos
        for col in range(1, 11):
            ws.cell(row=idx, column=col).border = borde

        # Insertar imagen
        if p.imagen:
            try:
                ruta_imagen = p.imagen.path
                if os.path.exists(ruta_imagen):
                    img = ExcelImage(ruta_imagen)
                    img.width = 100  # ancho mayor para mejor visualización
                    img.height = 100
                    celda_imagen = f'K{idx}'
                    ws.add_image(img, celda_imagen)
                    # Ajustar alto de fila para que la imagen quepa bien
                    ws.row_dimensions[idx].height = 80
            except Exception as e:
                print(f"Error cargando imagen para producto {p.nombre}: {e}")

            # Bordes para la celda de imagen
            ws.cell(row=idx, column=11).border = borde

    # Ajustar ancho columnas
    anchos = [20, 15, 20, 15, 15, 12, 12, 10, 18, 18, 18]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = ancho

    # Congelar fila de encabezado
    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
    wb.save(response)
    return response
    
@login_required
@cargo_requerido(['Admin', 'Gerente'])
def modificar_producto(request, id):
    producto = get_object_or_404(Producto, pk=id)

    if request.method == 'POST':
        # Agrega request.FILES para manejar archivos como imágenes
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            mensaje = "Producto modificado correctamente"
            return render(request, 'facturacion_cliente/producto/modificar_producto.html', {
                'form_Producto': form,
                'mensaje': mensaje
            })
        else:
            print(form.errors)  # Para depurar errores del formulario
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'facturacion_cliente/producto/modificar_producto.html', {
        'form_Producto': form
    })
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def eliminar_producto(request: HttpRequest, id: int) -> HttpResponse:
    producto = get_object_or_404(Producto, id=id)
    if request.method == "POST":
        producto.delete()
        return redirect("consultar_producto")
    return render(request, "facturacion_cliente/producto/eliminar_producto.html", {"producto": producto})

######################################################################
# Funciones simples para renderizar templates sin lógica adicional
def filtrar_descuentos(nombre=None, porcentaje=None):
    descuentos = Descuento.objects.all()
    if nombre:
        descuentos = descuentos.filter(nombre__icontains=nombre)
    if porcentaje is not None:
        descuentos = descuentos.filter(porcentaje=porcentaje)
    return descuentos








##########################-------------------DESCUEENTO-------------##########################################
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def crear_descuento(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = DescuentoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("consultar_descuento")
    else:
        form = DescuentoForm()
    
    return render(request, "facturacion_cliente/descuento/crear_descuento.html", {
        "form": form,
        "edit_mode": False
    })
@login_required
@cargo_requerido(['admin', 'Gerente', 'supervisor', 'cajero'])
def consultar_descuento(request: HttpRequest) -> HttpResponse:
    buscar_descuento_form = BuscarDescuentoForm(request.POST or None)

    descuentos = Descuento.objects.all()

    if request.method == "POST":
        if 'buscar_descuento' in request.POST and buscar_descuento_form.is_valid():
            producto = buscar_descuento_form.cleaned_data.get('producto', '').strip()
            descuento = buscar_descuento_form.cleaned_data.get('descuento', None)
            fecha = buscar_descuento_form.cleaned_data.get('fecha', None)

            if producto:
                descuentos = descuentos.filter(
                    Q(producto__nombre__icontains=producto) |
                    Q(producto__numero_serie__icontains=producto)
                )

            if descuento is not None:
                descuentos = descuentos.filter(descuento=descuento)

            if fecha:
                # Filtra por fecha (solo la parte de la fecha, sin hora)
                descuentos = descuentos.filter(fecha_inicio__date=fecha.date())

        elif 'exportar_descuento' in request.POST:
            return exportar_pdf_descuento(request)

    return render(request, "facturacion_cliente/descuento/consultar_descuento.html", {
        'buscador_descuento': buscar_descuento_form,
        'descuentos': descuentos
    })


@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def exportar_pdf_descuento(request: HttpRequest) -> HttpResponse:
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="listado_descuentos.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=inch / 2,
        leftMargin=inch / 2,
        topMargin=inch / 2,
        bottomMargin=inch / 2
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenteredTitle', alignment=TA_CENTER, fontSize=18, leading=22, spaceAfter=15, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='RightAlign', alignment=TA_RIGHT, fontSize=10))
    styles.add(ParagraphStyle(name='TableHeader', alignment=TA_CENTER, fontSize=11, fontName='Helvetica-Bold', textColor=colors.whitesmoke))
    styles.add(ParagraphStyle(name='TableCell', alignment=TA_LEFT, fontSize=10, leading=12))
    
    elementos = []

    # Título
    titulo = Paragraph("Listado de Descuentos", styles['CenteredTitle'])
    elementos.append(titulo)

    # Fecha de emisión
    fecha = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    elementos.append(Paragraph(f"Fecha de emisión: {fecha}", styles['RightAlign']))
    elementos.append(Spacer(1, 12))

    # Obtener descuentos
    descuentos = Descuento.objects.all()

    if not descuentos.exists():
        elementos.append(Paragraph("No hay descuentos registrados.", styles['Normal']))
    else:
        headings = ['ID', 'Producto', 'Descripción', 'Descuento', 'Fecha Inicio', 'Fecha Final', 'Estado']

        datos = []
        for d in descuentos:
            estado_texto = dict(Descuento.ESTADO_CHOICES).get(d.estado, "Desconocido") if hasattr(Descuento, 'ESTADO_CHOICES') else str(d.estado)
            datos.append([
                str(d.pk),
                getattr(d.producto, 'nombre', str(d.producto)),
                d.descripcion or '-',
                f"{d.descuento:.2f}",
                d.fecha_inicio.strftime("%d/%m/%Y"),
                d.fecha_final.strftime("%d/%m/%Y"),
                estado_texto
            ])

        # Construir tabla con encabezados y datos
        tabla = Table([headings] + datos, colWidths=[40, 100, 140, 70, 70, 70, 60])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2E86C1")),  # Azul oscuro para encabezado
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),

            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Descuento a la derecha
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # ID centrado

            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ]))
        elementos.append(tabla)

    # Pie de página con número de página
    def agregar_pie(canvas, doc):
        canvas.saveState()
        footer_text = f"Página {doc.page}"
        canvas.setFont('Helvetica', 8)
        width, height = A4
        canvas.drawRightString(width - inch / 2, 0.75 * inch, footer_text)
        canvas.restoreState()

    doc.build(elementos, onFirstPage=agregar_pie, onLaterPages=agregar_pie)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def modificar_descuento(request: HttpRequest, id: int) -> HttpResponse:
    descuento = get_object_or_404(Descuento, id=id)
    if request.method == "POST":
        form = DescuentoForm(request.POST, instance=descuento)
        if form.is_valid():
            form.save()
            return redirect("consultar_descuento")
    else:
        form = DescuentoForm(instance=descuento)
    
    return render(request, "facturacion_cliente/descuento/modificar_descuento.html", {
        "form": form,
        "descuento": descuento,
        "edit_mode": True
    })

@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def eliminar_descuento(request: HttpRequest, id: int) -> HttpResponse:
    descuento = get_object_or_404(Descuento, id=id)
    if request.method == "POST":
        descuento.delete()
        return redirect("consultar_descuento")
    return render(request, "facturacion_cliente/descuento/eliminar_descuento.html", {
        "descuento": descuento
    })
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def crear_proveedor(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = ProveedorForm(data=request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            return redirect("consultar_proveedor")
    else:
        form = ProveedorForm()
    context = {"form_Proveedor": form, "edit_mode": False}
    return render(request, "administrativo/proveedor/crear_proveedor.html", context)



@login_required
@cargo_requerido(['admin', 'Gerente', 'supervisor', 'cajero'])
def consultar_proveedor(request):
    # ✅ Si es una petición AJAX (desde Select2)
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        termino = request.GET.get('q', '').strip()
        proveedores = Proveedor.objects.filter(nombre__icontains=termino)[:10]
        resultados = [{'id': p.pk, 'text': p.nombre} for p in proveedores]
        return JsonResponse({'results': resultados})

    # ✅ Si es una petición normal (navegador)
    buscarProveedorForm = BuscarProveedorForm(request.POST or None)
    proveedores = Proveedor.objects.all()

    if request.method == "POST":
        if 'buscar_proveedor' in request.POST and buscarProveedorForm.is_valid():
            nombre = buscarProveedorForm.cleaned_data.get('nombre', '').strip()
            ruc = buscarProveedorForm.cleaned_data.get('ruc', '').strip()
            fecha = buscarProveedorForm.cleaned_data.get('fecha', None)

            filtros = Q()
            if nombre:
                filtros &= Q(nombre__icontains=nombre)
            if ruc:
                filtros &= Q(ruc__icontains=ruc)
            if fecha:
                filtros &= Q(fecha_creacion__lte=fecha)

            proveedores = Proveedor.objects.filter(filtros)

        elif 'exportar_proveedor' in request.POST:
            return exportar_pdf_proveedor(request)

    return render(request, "administrativo/proveedor/consultar_proveedor.html", {
        'buscador_Proveedor': buscarProveedorForm,
        'proveedores': proveedores
    })


@login_required
@cargo_requerido(['admin', 'Gerente', 'supervisor'])
def exportar_pdf_proveedor(request: HttpRequest) -> HttpResponse:
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="consultar_proveedor.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        rightMargin=inch / 2,
        leftMargin=inch / 2,
        topMargin=inch,
        bottomMargin=inch / 2,
        pagesize=A4
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='CenteredTitle', alignment=TA_CENTER, fontSize=16, spaceAfter=10))
    styles.add(ParagraphStyle(name='RightAlign', alignment=TA_RIGHT, fontSize=10))

    elementos = []

    # Encabezado
    elementos.append(Paragraph("Sistema de Gestión de Proveedores", styles['CenteredTitle']))
    elementos.append(Paragraph("Listado de Proveedores", styles['Title']))
    elementos.append(Paragraph(f"Fecha: {now().strftime('%d/%m/%Y %H:%M')}", styles['RightAlign']))
    elementos.append(Spacer(1, 12))

    # Filtro de datos si hay formulario
    buscarProveedorForm = BuscarProveedorForm(request.POST or None)
    if buscarProveedorForm.is_valid():
        nombre = buscarProveedorForm.cleaned_data.get('nombre', '')
        ruc = buscarProveedorForm.cleaned_data.get('ruc', '')
        fecha = buscarProveedorForm.cleaned_data.get('fecha')

        proveedores = Proveedor.objects.filter(
            Q(nombre__icontains=nombre) if nombre else Q(),
            Q(ruc__icontains=ruc) if ruc else Q(),
            Q(fecha_creacion__lte=fecha) if fecha else Q(),
        )
    else:
        proveedores = Proveedor.objects.all()

    # Cabecera de la tabla
    encabezado = [
        'ID', 'Nombre', 'Marca', 'RUC', 'Teléfono', 'Correo', 'Dirección', 'Estado', 'F. Creación'
    ]
    datos = [encabezado]

    for p in proveedores:
        datos.append([
            p.pk,
            p.nombre,
            p.marca,
            p.ruc,
            p.telefono,
            p.correo,
            p.direccion,
            "Activo" if p.estado == 1 else "Inactivo",
            p.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])

    tabla = Table(datos, colWidths=[40, 70, 60, 70, 60, 100, 80, 50, 80])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
    ]))

    elementos.append(tabla)
    doc.build(elementos)
    
    response.write(buffer.getvalue())
    buffer.close()
    return response
   
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def modificar_proveedor(request: HttpRequest, id: int) -> HttpResponse:
    proveedor = get_object_or_404(Proveedor, id=id)
    if request.method == "POST":
        form = ProveedorForm(data=request.POST, files=request.FILES, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect("consultar_proveedor")
    else:
        form = ProveedorForm(instance=proveedor)
    context = {"form": form, "proveedor": proveedor, "edit_mode": True}
    return render(request, "administrativo/proveedor/modificar_proveedor.html", context)

@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def eliminar_proveedor(request: HttpRequest, id: int) -> HttpResponse:
    proveedor = get_object_or_404(Proveedor, id=id)
    if request.method == "POST":
        proveedor.delete()
        return redirect("consultar_proveedor")
    return render(request, "administrativo/proveedor/eliminar_proveedor.html", {"proveedor": proveedor})

####################################### ------cotización -------------------    ############################
@login_required
@cargo_requerido(['admin', 'Gerente', 'cajero','supervisor' ])
def crear_cotizacion(request: HttpRequest) -> HttpResponse:
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        empleado_id = request.POST.get('empleado_id')
        sucursal_id = request.POST.get('sucursal_id')
        comentario = request.POST.get('comentario', '')
        productos_json = request.POST.get('productos_json')

        if not cliente_id or not productos_json:
            return HttpResponse("Datos incompletos.", status=400)

        try:
            productos = json.loads(productos_json)
        except json.JSONDecodeError:
            return HttpResponse("Error en el formato de productos.", status=400)

        try:
            with transaction.atomic():
                cotizacion = Cotizacion.objects.create(
                    cliente_id=cliente_id,
                    empleado_id=empleado_id if empleado_id else None,
                    sucursal_id=sucursal_id,
                    fecha_emision=timezone.now(),
                    comentario=comentario,
                    creacion_usuario=request.user.username,
                    modificacion_usuario=request.user.username,
                )

                ahora = timezone.now()

                for p in productos:
                    producto = Producto.objects.get(id=p['id'])
                    cantidad = int(p['cantidad'])
                    iva = Iva.objects.get(id=p['iva_id']) if p.get('iva_id') else None

                    # Obtener descuento activo (igual que en factura)
                    descuento_activo = Descuento.objects.filter(
                        producto=producto,
                        estado=1,
                        fecha_inicio__lte=ahora,
                        fecha_final__gte=ahora
                    ).order_by('-fecha_inicio').first()

                    precio_base = producto.precio
                    if descuento_activo and descuento_activo.descuento > 0:
                        descuento_decimal = Decimal(descuento_activo.descuento) / Decimal('100')
                        precio_desc = (precio_base * (Decimal('1') - descuento_decimal)).quantize(Decimal('0.01'))
                        descuento_total = Decimal(descuento_activo.descuento)
                    else:
                        precio_desc = precio_base
                        descuento_total = Decimal('0')

                    # Aquí NO modificamos stock, porque es cotización

                    Cotizacion_Detalle.objects.create(
                        cotizacion=cotizacion,
                        producto=producto,
                        cantidad_producto=cantidad,
                        precio_cotizado=precio_desc,
                        descuento_total=descuento_total,
                        iva=iva,
                        creacion_usuario=request.user.username,
                        modificacion_usuario=request.user.username
                    )

                totales = cotizacion.get_totales()
                cotizacion.comentario = (cotizacion.comentario or "") + f" | Subtotal: {totales['subtotal']}, IVA: {totales['iva_total']}, Total: {totales['total']}"
                cotizacion.save()

        except Exception as e:
            return HttpResponse(f"Error al procesar la cotización: {str(e)}", status=500)

        pdf_url = reverse('exportar_pdf_cotizacion', kwargs={'cotizacion_id': cotizacion.id})
        return JsonResponse({'success': True, 'pdf_url': pdf_url})

    # Si es GET: preparar datos igual que en factura pero para cotización

    ahora = timezone.now()
    productos_queryset = Producto.objects.all()
    productos = []

    for p in productos_queryset:
        descuento_activo = Descuento.objects.filter(
            producto=p,
            estado=1,
            fecha_inicio__lte=ahora,
            fecha_final__gte=ahora
        ).order_by('-fecha_inicio').first()

        productos.append({
            'id': p.pk,
            'nombre': p.nombre,
            'numero_serie': p.numero_serie,
            'precio': float(p.precio),
            'stock': p.stock,  # solo para mostrar, no modificar
            'descuento': float(descuento_activo.descuento) if descuento_activo else 0
        })

    clientes = list(Cliente.objects.values('id', 'nombre', 'cedula'))
    empleados = list(Empleado.objects.values('id', 'nombre', 'cargo'))
    sucursales = list(Sucursal.objects.values('id', 'nombre', 'direccion'))
    ivas = Iva.objects.all()

    context = {
        'productos_json': json.dumps(productos, ensure_ascii=False),
        'clientes_json': json.dumps(clientes, ensure_ascii=False),
        'empleados_json': json.dumps(empleados, ensure_ascii=False),
        'sucursales_json': json.dumps(sucursales, ensure_ascii=False),
        'ivas': ivas,
    }

    return render(request, 'facturacion_cliente/cotizacion/crear_cotizacion.html', context)

@login_required
@cargo_requerido(['admin', 'Gerente', 'cajero','supervisor' ])
def consultar_cotizacion(request, id=None):
    cotizacion = None
    if id is not None:
        cotizacion = get_object_or_404(Cotizacion, pk=id)

    cotizaciones = None
    buscar_form = BuscarCotizacionForm(request.POST or None)
    filtros = Q()

    if request.method == "POST":
        if 'buscar_cotizacion' in request.POST and buscar_form.is_valid():
            cliente = buscar_form.cleaned_data.get('cliente_nombre')
            desde = buscar_form.cleaned_data.get('desde')
            hasta = buscar_form.cleaned_data.get('hasta')

            if cliente:
                filtros &= (
                    Q(cliente__nombre__icontains=cliente) |
                    Q(cliente__cedula__icontains=cliente)
                )
            if desde and hasta:
                filtros &= Q(fecha_emision__range=(desde, hasta))
            elif desde:
                filtros &= Q(fecha_emision__gte=desde)
            elif hasta:
                filtros &= Q(fecha_emision__lte=hasta)

            cotizaciones = Cotizacion.objects.filter(filtros).order_by('-fecha_emision')

            for c in cotizaciones:
                totales = c.get_totales()
                c.total_calculado = totales['total']  # type: ignore

        elif 'exportar_pdf_cotizacion' in request.POST:
            if cotizacion and cotizacion.id:
                return exportar_pdf_cotizacion(request, cotizacion.id)
            else:
                messages.error(request, "No se ha seleccionado una cotización para exportar.")

    context = {
        'buscador_cotizacion': buscar_form,
        'cotizaciones': cotizaciones,
        'cotizacion': cotizacion,
    }
    return render(request, 'facturacion_cliente/cotizacion/consultar_cotizacion.html', context)


@login_required
@cargo_requerido(['admin', 'Gerente', 'cajero', 'supervisor'])
def exportar_pdf_cotizacion(request, cotizacion_id):
    try:
        cotizacion = Cotizacion.objects.get(pk=cotizacion_id)
    except Cotizacion.DoesNotExist:
        return HttpResponse("Cotización no encontrada.", status=404)

    detalles = cotizacion.detalles.all()  # type: ignore
    sucursal = cotizacion.sucursal

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 50

    def dibujar_encabezado():
        nonlocal y
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y, f"Reporte de Cotización #{cotizacion.numero_cotizacion or cotizacion.id}")
        y -= 30

        p.setFont("Helvetica", 10)
        p.drawString(50, y, f"Sucursal: {sucursal.nombre}")
        y -= 20
        p.drawString(50, y, f"Fecha emisión: {cotizacion.fecha_emision.strftime('%d/%m/%Y')}")
        y -= 20
        p.drawString(50, y, f"Cliente: {cotizacion.cliente.nombre}")
        y -= 30

        # Encabezado de la tabla
        p.setFont("Helvetica-Bold", 9)
        p.drawString(50, y, "Producto")
        p.drawString(160, y, "Cant.")
        p.drawString(200, y, "P. Unit")
        p.drawString(270, y, "Desc.")
        p.drawString(320, y, "IVA %")
        p.drawString(380, y, "Subtotal")
        p.drawString(470, y, "Total")
        y -= 20

    dibujar_encabezado()

    total_general = Decimal('0.00')

    for d in detalles:
        producto = d.producto.nombre
        cantidad = d.cantidad_producto
        precio = d.precio_cotizado
        descuento = d.descuento_total or Decimal('0.00')
        iva_porcentaje = d.iva.porcentaje if d.iva else Decimal('0.00')

        # Cálculos
        subtotal = cantidad * precio
        subtotal_desc = subtotal 
        iva_valor = subtotal_desc * (iva_porcentaje / Decimal('100'))
        total = subtotal_desc + iva_valor
        total_general += total

        # Ajuste para nombres largos
        p.setFont("Helvetica", 9)
        lineas = simpleSplit(producto, "Helvetica", 9, 100)
        y_inicial = y
        for linea in lineas:
            p.drawString(50, y, linea)
            y -= 12

        # Otros campos en la primera línea
        p.drawString(160, y_inicial, str(cantidad))
        p.drawString(200, y_inicial, f"${precio:.2f}")
        p.drawString(270, y_inicial, f"${descuento:.2f}")
        p.drawString(320, y_inicial, f"{iva_porcentaje:.0f}%")
        p.drawString(380, y_inicial, f"${subtotal_desc:.2f}")
        p.drawString(470, y_inicial, f"${total:.2f}")

        y -= 10

        if y < 100:
            p.showPage()
            y = height - 50
            dibujar_encabezado()

    # Totales
    if y < 100:
        p.showPage()
        y = height - 50
    p.setFont("Helvetica-Bold", 11)
    y -= 20
    p.drawString(370, y, "TOTAL GENERAL:")
    p.drawString(470, y, f"${total_general:.2f}")

    # Finalizar PDF
    p.showPage()
    p.save()
    buffer.seek(0)

    filename = f"Cotizacion_{cotizacion.numero_cotizacion or cotizacion.id}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response



@login_required
@cargo_requerido(['admin', 'Gerente','supervisor','cajero'])
def modificar_cotizacion(request: HttpRequest, id: int) -> HttpResponse:
    cotizacion = get_object_or_404(Cotizacion, id=id)
    if request.method == "POST":
        form = CotizacionForm(request.POST, instance=cotizacion)
        formset = CotizacionDetalleFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            cotizacion = form.save(commit=False)
            cotizacion.creacion_usuario = request.user.username
            cotizacion.save()
            for detalle_form in formset:
                if detalle_form.cleaned_data and not detalle_form.cleaned_data.get('DELETE', False):
                    detalle = detalle_form.save(commit=False)
                    detalle.cotizacion = cotizacion
                    detalle.save()
            return redirect("consultar_cotizacion")
    else:
        form = CotizacionForm(instance=cotizacion)
        formset = CotizacionDetalleFormSet()
    context = {"form": form, "formset": formset, "cotizacion": cotizacion, "edit_mode": True}
    return render(request, "facturacion_cliente/cotizacion/modificar_cotizacion.html")


@login_required
@cargo_requerido(['admin', 'Gerente','supervisor','cajero'])
def eliminar_cotizacion(request, cotizacion_id):
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)

    if request.method == "POST":
        cotizacion.delete()
        return redirect("consultar_cotizacion")

    # Enviar cotización al template para mostrar info y confirmar
    context = {
        'cotizacion': cotizacion
    }
    return render(request, "facturacion_cliente/cotizacion/eliminar_cotizacion.html", context)



########################################### Cotización Detalle ########################################
# Funciones para manejar cotización detalle\


@login_required
@cargo_requerido(['admin', 'Gerente','supervisor','cajero'])
def crear_cotizacion_detalle(request):
    if request.method == "POST":
        form = CotizacionDetalleForm(request.POST)
        if form.is_valid():
            cotizacion_detalle = form.save(commit=False)
            cotizacion_detalle.creacion_usuario = request.user.username
            cotizacion_detalle.modificacion_usuario = request.user.username
            cotizacion_detalle.save()
            messages.success(request, "Detalle de cotización creado correctamente.")
            return redirect("consultar_cotizacion_detalle")  # Cambia esto por la vista donde quieres redirigir
        else:
            messages.error(request, "Por favor, corrige los errores en el formulario.")
    else:
        form = CotizacionDetalleForm()
    
    context = {
        'form': form,
    }
    return render(request, "facturacion_cliente/cotizacion_detalle/crear_cotizacion_detalle.html", context)

@login_required
@cargo_requerido(['admin', 'Gerente','supervisor','cajero'])
def consultar_cotizacion_detalle(request):
    detalles = None
    buscar_form = BuscarCotizacionDetalleForm()

    if request.method == "POST":
        if 'buscar_cotizacion_detalle' in request.POST:
            buscar_form = BuscarCotizacionDetalleForm(request.POST)
            if buscar_form.is_valid():
                producto_nombre = buscar_form.cleaned_data.get('producto_nombre')
                cotizacion_numero = buscar_form.cleaned_data.get('cotizacion_numero')
                desde = buscar_form.cleaned_data.get('desde')
                hasta = buscar_form.cleaned_data.get('hasta')

                filtros = Q()
                if producto_nombre:
                    filtros &= Q(producto__nombre__icontains=producto_nombre)
                if cotizacion_numero:
                    filtros &= Q(cotizacion__numero_cotizacion__icontains=cotizacion_numero)
                if desde and hasta:
                    filtros &= Q(fecha_creacion__range=(desde, hasta))
                elif desde:
                    filtros &= Q(fecha_creacion__gte=desde)
                elif hasta:
                    filtros &= Q(fecha_creacion__lte=hasta)

                detalles = Cotizacion_Detalle.objects.filter(filtros).order_by('-fecha_creacion')

    context = {
        'buscador_detalle': buscar_form,
        'detalles': detalles
    }
    return render(request, "facturacion_cliente/cotizacion_detalle/consultar_cotizacion_detalle.html", context)
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor','cajero'])
def eliminar_cotizacion_detalle(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        cotizacion_detalle_id = request.POST.get('cotizacion_detalle_id')
        if cotizacion_detalle_id:
            cotizacion_detalle = get_object_or_404(Cotizacion_Detalle, id=cotizacion_detalle_id)
            cotizacion_detalle.delete()
            messages.success(request, "Detalle de cotización eliminado correctamente.")
            # Aquí redirige a la vista que muestra los detalles de la cotización relacionada
            return redirect("consultar_cotizacion_detalle")  
        else:
            messages.error(request, "No se proporcionó el ID del detalle.")
            return redirect("consultar_cotizacion_detalle")

    # Si no es POST, puedes mostrar una confirmación o redirigir
    return render(request, "facturacion_cliente/cotizacion_detalle/eliminar_cotizacion_detalle.html")
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor','cajero'])
def modificar_cotizacion_detalle(request: HttpRequest, id: int) -> HttpResponse:
    cotizacion_detalle = get_object_or_404(Cotizacion_Detalle, id=id)
    
    if request.method == "POST":
        form = CotizacionDetalleForm(request.POST, instance=cotizacion_detalle)
        if form.is_valid():
            detalle_modificado = form.save(commit=False)
            detalle_modificado.modificacion_usuario = request.user.username
            detalle_modificado.save()
            messages.success(request, "Detalle de cotización modificado correctamente.")
            return redirect("consultar_cotizacion_detalle")
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = CotizacionDetalleForm(instance=cotizacion_detalle)
    
    return render(request, "facturacion_cliente/cotizacion_detalle/modificar_cotizacion_detalle.html", {
        'form': form,
        'cotizacion_detalle': cotizacion_detalle,
    })




######################################## Facturación Cliente ########################################



# Tus imports y modelos aquí (Factura, Producto, Iva, Factura_Detalle, etc.)
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor','cajero'])
def crear_factura(request: HttpRequest) -> HttpResponse:
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        empleado_id = request.POST.get('empleado_id')
        sucursal_id = request.POST.get('sucursal_id')
        tipo_pago = request.POST.get('tipo_pago')
        comentario = request.POST.get('comentario', '')
        productos_json = request.POST.get('productos_json')

        if not cliente_id or not productos_json:
            return HttpResponse("Datos incompletos.", status=400)

        try:
            productos = json.loads(productos_json)
        except json.JSONDecodeError:
            return HttpResponse("Error en el formato de productos.", status=400)

        try:
            with transaction.atomic():
                factura = Factura.objects.create(
                    cliente_id=cliente_id,
                    empleado_id=empleado_id if empleado_id else None,
                    sucursal_id=sucursal_id,
                    tipo_pago=tipo_pago,
                    fecha_emision=timezone.now(),
                    comentario=comentario,
                    creacion_usuario=request.user.username,
                    modificacion_usuario=request.user.username,
                )

                ahora = timezone.now()

                for p in productos:
                    producto = Producto.objects.get(id=p['id'])
                    cantidad = int(p['cantidad'])
                    iva = Iva.objects.get(id=p['iva_id']) if p.get('iva_id') else None

                    if producto.stock < cantidad:
                        return HttpResponse(f"Stock insuficiente para {producto.nombre}", status=400)

                    descuento_total = Decimal(p.get('descuento_pct', 0))  # viene desde el frontend
                    precio_base = producto.precio
                    precio_desc = precio_base

                    if descuento_total > 0:
                        descuento_decimal = descuento_total / Decimal('100')
                        precio_desc = (precio_base * (Decimal('1') - descuento_decimal)).quantize(Decimal('0.01'))

                    producto.stock -= cantidad
                    producto.save()

                    Factura_Detalle.objects.create(
                        factura=factura,
                        producto=producto,
                        cantidad_producto=cantidad,
                        precio_factura=precio_desc,
                        descuento_total=descuento_total,
                        iva=iva,
                        creacion_usuario=request.user.username,
                        modificacion_usuario=request.user.username
                    )

                totales = factura.get_totales
                factura.comentario = (factura.comentario or "") + f" | Subtotal: {totales['subtotal']}, IVA: {totales['iva_total']}, Total: {totales['total']}"
                factura.save()

        except Exception as e:
            return HttpResponse(f"Error al procesar la factura: {str(e)}", status=500)

        pdf_factura = reverse('exportar_pdf_factura', kwargs={'factura_id': factura.id})
        return JsonResponse({'success': True, 'pdf_factura': pdf_factura})



    # Si es GET: preparar datos
    ahora = timezone.now()
    productos_queryset = Producto.objects.all()
    productos = []

    for p in productos_queryset:
        descuento_activo = Descuento.objects.filter(
            producto=p,
            estado=1,
            fecha_inicio__lte=ahora,
            fecha_final__gte=ahora
        ).order_by('-fecha_inicio').first()

        productos.append({
            'id': p.pk,
            'nombre': p.nombre,
            'numero_serie': p.numero_serie,
            'precio': float(p.precio),
            'stock': p.stock,
            'descuento': float(descuento_activo.descuento) if descuento_activo else 0
        })

    clientes = list(Cliente.objects.values('id', 'nombre', 'cedula'))
    empleados = list(Empleado.objects.values('id', 'nombre', 'cargo'))
    sucursales = list(Sucursal.objects.values('id', 'nombre', 'direccion'))
    ivas = Iva.objects.all()

    context = {
        'productos_json': json.dumps(productos, ensure_ascii=False),
        'clientes_json': json.dumps(clientes, ensure_ascii=False),
        'empleados_json': json.dumps(empleados, ensure_ascii=False),
        'sucursales_json': json.dumps(sucursales, ensure_ascii=False),
        'ivas': ivas,
    }

    return render(request, 'facturacion_cliente/factura/crear_factura.html', context)


@login_required
@cargo_requerido(['admin', 'Gerente','supervisor','cajero'])
def crear_factura_detalle(request):
    if request.method == 'POST':
        factura_form = FacturaForm(request.POST)
        detalle_formset = FacturaDetalleFormSet(request.POST)
        if factura_form.is_valid() and detalle_formset.is_valid():
            with transaction.atomic():
                factura = factura_form.save()
                detalles = detalle_formset.save(commit=False)
                for detalle in detalles:
                    detalle.factura = factura
                    detalle.save()
            return redirect('consultar_factura')  # Cambiado a nombre de URL
    else:
        factura_form = FacturaForm()
        detalle_formset = FacturaDetalleFormSet()

    context = {
        'factura_form': factura_form,
        'detalle_formset': detalle_formset,
    }
    return render(request, "facturacion_cliente/factura/consultar_factura.html", context)

@login_required
@cargo_requerido(['admin', 'Gerente','supervisor','cajero'])
def consultar_factura(request, id=None):
    factura = None
    if id is not None:
        factura = get_object_or_404(Factura, pk=id)
    
    facturas = None
    buscar_form = BuscarFacturaForm(request.POST or None)
    filtros = Q()

    if request.method == "POST":
        if 'buscar_factura' in request.POST and buscar_form.is_valid():
            cedula = buscar_form.cleaned_data.get('cliente_cedula')
            desde = buscar_form.cleaned_data.get('desde')
            hasta = buscar_form.cleaned_data.get('hasta')

            if cedula:
                filtros &= Q(cliente__cedula__icontains=cedula)
            if desde and hasta:
                filtros &= Q(fecha_creacion__range=(desde, hasta))
            elif desde:
                filtros &= Q(fecha_creacion__gte=desde)
            elif hasta:
                filtros &= Q(fecha_creacion__lte=hasta)

            facturas = Factura.objects.filter(filtros).order_by('-fecha_creacion')

            # Calcular total para cada factura
            for f in facturas:
                total = f.detalles.aggregate(total=Sum('total_factura_valor'))['total'] or 0  # type: ignore
                f.total_calculado = total  # type: ignore

        elif 'exportar_factura' in request.POST:
            if factura and factura.id:
                return exportar_pdf_factura(request, factura.id)
            

    context = {
        'buscador_factura': buscar_form,
        'facturas': facturas,
        'factura': factura,
    }
    return render(request, 'facturacion_cliente/factura/consultar_factura.html', context)



@login_required
@cargo_requerido(['admin', 'Gerente', 'cajero', 'supervisor'])
def exportar_pdf_factura(request, factura_id):
    if not factura_id:
        return HttpResponse("ID de factura no proporcionado.", status=400)

    try:
        factura = Factura.objects.select_related('cliente', 'sucursal').get(pk=factura_id)
    except Factura.DoesNotExist:
        return HttpResponse("Factura no encontrada.", status=404)

    cliente = factura.cliente
    sucursal = factura.sucursal
    detalles = factura.detalles.filter(estado=1)  # type: ignore

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    MARGIN = 50
    TOP = height - MARGIN
    RIGHT_X = width - 250

    # === ENCABEZADO ===
    p.setFont("Helvetica-Bold", 10)
    p.drawString(MARGIN, TOP - 15, f"R.U.C.: {sucursal.ruc}")
    p.setFont("Helvetica", 9)
    p.drawString(MARGIN, TOP - 30, sucursal.nombre.upper())
    p.drawString(MARGIN, TOP - 45, f"Local: {sucursal.local}")
    p.drawString(MARGIN, TOP - 60, f"Dirección: {sucursal.direccion}")
    p.drawString(MARGIN, TOP - 75, f"Teléfono: {sucursal.telefono}")
    p.drawString(MARGIN, TOP - 90, f"Correo: {sucursal.correo}")
    p.drawString(MARGIN, TOP - 105, "Obligado a llevar contabilidad: SI")
    p.drawString(MARGIN, TOP - 120, "Emisión: NORMAL")

    # Datos de factura
    p.setFont("Helvetica-Bold", 10)
    p.drawString(RIGHT_X, TOP - 15, f"FACTURA No.: {factura.numero_factura or 'N/A'}")
    p.setFont("Helvetica", 9)
    p.drawString(RIGHT_X, TOP - 30, "AUTORIZACIÓN:")
    p.setFont("Helvetica", 7)
    p.drawString(RIGHT_X, TOP - 45, getattr(factura, 'numero_autorizacion', '000000000000000'))
    p.setFont("Helvetica", 9)
    p.drawString(RIGHT_X, TOP - 60, f"Fecha Emisión: {factura.fecha_emision.strftime('%d/%m/%Y')}")
    

    p.line(MARGIN, TOP - 135, width - MARGIN, TOP - 135)

    # === CLIENTE ===
    p.setFont("Helvetica-Bold", 10)
    p.drawString(MARGIN, TOP - 150, "CLIENTE:")
    p.setFont("Helvetica", 9)
    p.drawString(MARGIN + 70, TOP - 150, f"{cliente.nombre} {getattr(cliente, 'apellido', '')}   C.I.: {cliente.cedula}")
    p.drawString(MARGIN + 70, TOP - 165, f"Dirección: {cliente.direccion or 'No disponible'}")
    p.drawString(MARGIN + 70, TOP - 180, f"Teléfono: {cliente.telefono or 'N/A'}")
    p.drawString(MARGIN + 70, TOP - 195, f"Email: {cliente.correo or 'N/A'}")

    # === DETALLES DE PRODUCTOS ===
    data = [["Producto", "Cantidad", "Precio Unitario", "Descuento %", "Subtotal", "IVA", "Total"]]
    subtotal_general = iva_general = total_general = Decimal('0.00')

    for detalle in detalles:
        cantidad = Decimal(detalle.cantidad_producto)
        precio_unitario = detalle.precio_factura or Decimal('0.00')
        descuento = detalle.descuento_total or Decimal('0.00')
        iva_pct = Decimal(detalle.iva.porcentaje) if detalle.iva else Decimal('0.00')

        subtotal = (precio_unitario * cantidad) * (1 - descuento / 100)
        iva = subtotal * (iva_pct / 100)
        total = subtotal + iva

        subtotal_general += subtotal
        iva_general += iva
        total_general += total

        data.append([
            str(detalle.producto),
            str(cantidad),
            f"{precio_unitario:.2f}",
            f"{descuento:.2f}%",
            f"{subtotal:.2f}",
            f"{iva:.2f}",
            f"{total:.2f}"
        ])

    table = Table(data, colWidths=[140, 60, 70, 70, 70, 60, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#dbe5f1")),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
    ]))

    table_width, table_height = table.wrap(0, 0)
    y_table = TOP - 230 - table_height
    if y_table < 100:
        p.showPage()
        y_table = height - 100

    table.drawOn(p, MARGIN, y_table)

    # === TOTALES ===
    y_totales = y_table - 40
    p.setFont("Helvetica-Bold", 10)
    p.setFillColor(colors.black)
    p.drawRightString(width - 60, y_totales, f"Subtotal: {subtotal_general:.2f}")
    p.drawRightString(width - 60, y_totales - 15, f"IVA Total: {iva_general:.2f}")
    p.setFont("Helvetica-Bold", 12)
    p.setFillColor(colors.HexColor("#003366"))
    p.drawRightString(width - 60, y_totales - 35, f"TOTAL A PAGAR: {total_general:.2f}")

    # === FOOTER ===
    p.setFont("Helvetica-Oblique", 8)
    p.setFillColor(colors.black)
    p.drawCentredString(width / 2, 30, "Gracias por su compra. Documento generado electrónicamente.")

    p.showPage()
    p.save()
    buffer.seek(0)

    return HttpResponse(
        buffer,
        content_type='application/pdf',
        headers={'Content-Disposition': f'inline; filename="factura_{factura.id}.pdf"'}
    )


@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])
def modificar_factura(request: HttpRequest, id: int) -> HttpResponse:
    factura = get_object_or_404(Factura, id=id)
    if request.method == "POST":
        form = FacturaForm(request.POST, instance=factura)
        if form.is_valid():
            factura_modificada = form.save(commit=False)
            factura_modificada.modificacion_usuario = request.user.username
            factura_modificada.save()
            messages.success(request, "Factura modificada correctamente.")
            return redirect('consultar_factura')
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = FacturaForm(instance=factura)

    context = {
        'form': form,
        'factura': factura,
    }
    return render(request, "facturacion_cliente/factura/modificar_factura.html", context)
@login_required
@cargo_requerido(['Admin', 'Gerente'])
def anular_factura(request, id):
    factura = get_object_or_404(Factura, pk=id)

    if request.method == 'POST':
        factura.estado = 0
        factura.save()
        return redirect('consultar_factura')

    return render(request, 'factura/anular.html', {'factura': factura})
@login_required
@cargo_requerido(['admin', 'Gerente','supervisor'])


def eliminar_factura(request, id):
    factura = get_object_or_404(Factura, id=id)

    if request.method == "POST":
        detalles = Factura_Detalle.objects.filter(factura=factura)

        for detalle in detalles:
            producto = detalle.producto
            producto.stock += detalle.cantidad_producto
            producto.save()

        # Primero eliminamos los detalles y luego la factura
        detalles.delete()
        factura.delete()

        messages.success(request, "Factura eliminada correctamente y stock restaurado.")
        return redirect('consultar_factura')

    context = {
        'factura': factura,
    }
    return render(request, "facturacion_cliente/factura/eliminar_factura.html", context)


######################################### Proveedor#########################################





##############################################################################################################
##############################################################################################################
#############################################################################################################
###############################################################################################################
########################################## Pagos #########################################

# import requests
# from django.conf import settings
# from django.http import JsonResponse

# def iniciar_pago_payphone(request):
#     url = "https://pay.payphonetodoesposible.com/api/Sale"
#     headers = {
#         "Authorization": f"Bearer {settings.PAYPHONE_API_KEY}",
#         "Content-Type": "application/json"
#     }

#     data = {
#         "amount": 1500,  # Valor en centavos: 15.00 USD
#         "amountWithoutTax": 1300,
#         "tax": 200,
#         "clientTransactionId": "TRX123ABC",
#         "phoneNumber": "0999999999",  # Obligatorio para PayPhone
#         "email": "cliente@example.com",
#         "storeId": settings.PAYPHONE_STORE_ID,
#         "paymentId": 0  # 0: pago con tarjeta / 1: pago con saldo PayPhone
#     }

#     response = requests.post(url, headers=headers, json=data)
#     return JsonResponse(response.json())



# @login_required
# @cargo_requerido(['admin', 'Gerente', 'supervisor', 'cajero'])
# def crear_factura(request: HttpRequest) -> HttpResponse:
#     if request.method == 'POST':
#         cliente_id = request.POST.get('cliente_id')
#         empleado_id = request.POST.get('empleado_id')
#         sucursal_id = request.POST.get('sucursal_id')
#         tipo_pago = request.POST.get('tipo_pago')
#         comentario = request.POST.get('comentario', '')
#         productos_json = request.POST.get('productos_json')
#         transaccion_id = request.POST.get('transaccion_id')  # <-- se recibe desde el frontend

#         if not cliente_id or not productos_json:
#             return HttpResponse("Datos incompletos.", status=400)

#         try:
#             productos = json.loads(productos_json)
#         except json.JSONDecodeError:
#             return HttpResponse("Error en el formato de productos.", status=400)

#         try:
#             with transaction.atomic():
#                 factura = Factura.objects.create(
#                     cliente_id=cliente_id,
#                     empleado_id=empleado_id if empleado_id else None,
#                     sucursal_id=sucursal_id,
#                     tipo_pago=tipo_pago,
#                     fecha_emision=timezone.now(),
#                     comentario=comentario,
#                     creacion_usuario=request.user.username,
#                     modificacion_usuario=request.user.username,
#                 )

#                 ahora = timezone.now()

#                 for p in productos:
#                     producto = Producto.objects.get(id=p['id'])
#                     cantidad = int(p['cantidad'])
#                     iva = Iva.objects.get(id=p['iva_id']) if p.get('iva_id') else None

#                     if producto.stock < cantidad:
#                         return HttpResponse(f"Stock insuficiente para {producto.nombre}", status=400)

#                     descuento_total = Decimal(p.get('descuento_pct', 0))
#                     precio_base = producto.precio
#                     precio_desc = precio_base

#                     if descuento_total > 0:
#                         descuento_decimal = descuento_total / Decimal('100')
#                         precio_desc = (precio_base * (Decimal('1') - descuento_decimal)).quantize(Decimal('0.01'))

#                     producto.stock -= cantidad
#                     producto.save()

#                     Factura_Detalle.objects.create(
#                         factura=factura,
#                         producto=producto,
#                         cantidad_producto=cantidad,
#                         precio_factura=precio_desc,
#                         descuento_total=descuento_total,
#                         iva=iva,
#                         creacion_usuario=request.user.username,
#                         modificacion_usuario=request.user.username
#                     )

#                 totales = factura.get_totales
#                 factura.comentario = (factura.comentario or "") + f" | Subtotal: {totales['subtotal']}, IVA: {totales['iva_total']}, Total: {totales['total']}"
#                 factura.save()

#                 # --- REGISTRAR PAGO TARJETA ---
#                 if tipo_pago == 'tarjeta':
#                     if not transaccion_id:
#                         return HttpResponse("No se recibió transacción del pago con tarjeta.", status=400)

#                     Pago.objects.create(
#                         factura=factura,
#                         metodo='tarjeta',
#                         estado='pagado',
#                         transaccion_id=transaccion_id,
#                         monto=totales['total']
#                     )

#         except Exception as e:
#             return HttpResponse(f"Error al procesar la factura: {str(e)}", status=500)

#         pdf_factura = reverse('exportar_pdf_factura', kwargs={'factura_id': factura.id})
#         return JsonResponse({'success': True, 'pdf_factura': pdf_factura})

#     # GET → preparar datos
#     ahora = timezone.now()
#     productos_queryset = Producto.objects.all()
#     productos = []

#     for p in productos_queryset:
#         descuento_activo = Descuento.objects.filter(
#             producto=p,
#             estado=1,
#             fecha_inicio__lte=ahora,
#             fecha_final__gte=ahora
#         ).order_by('-fecha_inicio').first()

#         productos.append({
#             'id': p.pk,
#             'nombre': p.nombre,
#             'numero_serie': p.numero_serie,
#             'precio': float(p.precio),
#             'stock': p.stock,
#             'descuento': float(descuento_activo.descuento) if descuento_activo else 0
#         })

#     clientes = list(Cliente.objects.values('id', 'nombre', 'cedula'))
#     empleados = list(Empleado.objects.values('id', 'nombre', 'cargo'))
#     sucursales = list(Sucursal.objects.values('id', 'nombre', 'direccion'))
#     ivas = Iva.objects.all()

#     context = {
#         'productos_json': json.dumps(productos, ensure_ascii=False),
#         'clientes_json': json.dumps(clientes, ensure_ascii=False),
#         'empleados_json': json.dumps(empleados, ensure_ascii=False),
#         'sucursales_json': json.dumps(sucursales, ensure_ascii=False),
#         'ivas': ivas,
#     }

#     return render(request, 'facturacion_cliente/factura/crear_factura.html', context)

@login_required
@cargo_requerido(['admin', 'Gerente', 'supervisor', 'cajero'])
def crear_compra(request):
    if request.method == 'POST':
        form = CompraForm(request.POST)
        if form.is_valid():
            compra = form.save(commit=False)
            compra.estado = 1  # por ejemplo, 1 para activo
            compra.save()
            messages.success(request, "Compra registrada exitosamente.")
            return redirect('consultar_compra')
    else:
        form = CompraForm()

    return render(request, 'administrativo/compra/crear_compra.html', {'form': form})






@login_required
@cargo_requerido(['admin', 'Gerente', 'supervisor', 'cajero'])
def consultar_compra(request):
    compras = Compra.objects.all()
    
    if request.method == 'POST':
        buscador_compra = BuscadorCompraForm(request.POST)
        if buscador_compra.is_valid():
            proveedor = buscador_compra.cleaned_data.get('proveedor')
            numero_factura = buscador_compra.cleaned_data.get('numero_factura')
            fecha_inicio = buscador_compra.cleaned_data.get('fecha_inicio')
            fecha_fin = buscador_compra.cleaned_data.get('fecha_fin')

            if proveedor:
                compras = compras.filter(proveedor=proveedor)
            if numero_factura:
                compras = compras.filter(numero_factura__icontains=numero_factura)
            if fecha_inicio:
                compras = compras.filter(fecha_compra__gte=fecha_inicio)
            if fecha_fin:
                compras = compras.filter(fecha_compra__lte=fecha_fin)
    else:
        buscador_compra = BuscadorCompraForm()

    return render(request, 'administrativo/compra/consultar_compra.html', {
        'compras': compras,
        'buscador_compra': buscador_compra,
    })


@login_required
@cargo_requerido(['admin', 'Gerente', 'supervisor'])
def modificar_compra(request, id):
    compra = get_object_or_404(Compra, id=id)

    if request.method == 'POST':
        form = CompraForm(request.POST, instance=compra)
        if form.is_valid():
            form.save()
            messages.success(request, 'Compra modificada exitosamente.')
            return redirect('consultar_compra')
    else:
        form = CompraForm(instance=compra)

    context = {
        'form': form,
        'compra': compra,
        'edit_mode': True,
    }
    return render(request, 'administrativo/compra/modificar_compra.html', context)
 
@login_required
@cargo_requerido(['admin', 'Gerente', 'supervisor'])
def anular_compra(request, id):
    compra = get_object_or_404(Compra, id=id)
    
    if request.method == 'POST':
        compra.estado = False  # o 0
        compra.save()
        messages.success(request, 'La compra ha sido anulada correctamente.')
        return redirect('consultar_compra')

    return render(request, 'administrativo/compra/anular_compra.html', {'compra': compra})

@login_required
@cargo_requerido(['admin', 'Gerente'])
def eliminar_compra(request, id):
    compra = get_object_or_404(Compra, id=id)
    compra.delete()
    messages.error(request, 'Compra eliminada permanentemente.')
    return redirect('consultar_compra')



def ajustar_y_centrar_texto(canvas, texto, x_centro, y, max_width, font_name="Helvetica", font_size=10, min_size=6):
    texto = str(texto)
    actual_size = font_size
    while stringWidth(texto, font_name, actual_size) > max_width and actual_size > min_size:
        actual_size -= 0.5
    canvas.setFont(font_name, actual_size)
    canvas.drawCentredString(x_centro, y, texto)
    canvas.setFont(font_name, font_size)  # Restaurar tamaño original

@login_required
@cargo_requerido(['admin', 'Gerente', 'supervisor'])
def exportar_pdf_compra(request):
    proveedor_id = request.GET.get('proveedor')
    numero_factura = request.GET.get('numero_factura')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    compras = Compra.objects.all()

    if proveedor_id:
        compras = compras.filter(proveedor_id=proveedor_id)
    if numero_factura:
        compras = compras.filter(numero_factura__icontains=numero_factura)
    if fecha_inicio:
        compras = compras.filter(fecha_compra__gte=fecha_inicio)
    if fecha_fin:
        compras = compras.filter(fecha_compra__lte=fecha_fin)

    total_iva = compras.aggregate(total_iva=Sum('iva'))['total_iva'] or 0

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="compras_filtradas.pdf"'

    c = canvas.Canvas(response, pagesize=letter) # type: ignore
    width, height = letter
    margin = 0.75 * inch
    y = height - margin

    # Título centrado
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width / 2, y, "Reporte de Compras Filtradas")
    y -= 0.4 * inch

    # Fecha y filtros
    c.setFont("Helvetica", 10)
    fecha_hoy = datetime.now().strftime('%d/%m/%Y %H:%M')
    c.drawString(margin, y, f"Fecha de generación: {fecha_hoy}")
    y -= 0.25 * inch

    filtros_aplicados = []
    if proveedor_id:
        proveedor_nombre = compras.first().proveedor.nombre if compras.exists() else "N/A" # type: ignore
        filtros_aplicados.append(f"Proveedor: {proveedor_nombre}")
    if numero_factura:
        filtros_aplicados.append(f"N° Factura contiene: '{numero_factura}'")
    if fecha_inicio:
        filtros_aplicados.append(f"Desde: {fecha_inicio}")
    if fecha_fin:
        filtros_aplicados.append(f"Hasta: {fecha_fin}")

    filtros_text = "Filtros aplicados: " + (", ".join(filtros_aplicados) if filtros_aplicados else "Ninguno")
    c.drawString(margin, y, filtros_text)
    y -= 0.5 * inch

    # Definir columnas
    headers = ["N° Factura", "Proveedor", "Fecha", "Subtotal", "IVA", "Total"]
    col_widths = [1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch]
    col_positions = [margin]
    for w in col_widths[:-1]:
        col_positions.append(col_positions[-1] + w)

    # Encabezado con fondo gris
    header_height = 0.3 * inch
    c.setFillColorRGB(0.9, 0.9, 0.9)
    c.rect(margin - 3, y - header_height + 5, sum(col_widths) + 6, header_height, fill=1, stroke=0)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 12)
    for i, header in enumerate(headers):
        x_center = col_positions[i] + col_widths[i] / 2
        c.drawCentredString(x_center, y - 0.2*inch, header)
    y -= header_height + 2

    # Línea bajo encabezado
    c.setLineWidth(1)
    c.line(margin - 3, y + 5, margin - 3 + sum(col_widths) + 6, y + 5)
    y -= 0.1 * inch

    c.setFont("Helvetica", 10)
    line_height = 0.25 * inch

    for compra in compras:
        if y < margin + line_height + 50:
            c.showPage()
            y = height - margin

            # Repetir encabezados
            c.setFillColorRGB(0.9, 0.9, 0.9)
            c.rect(margin - 3, y - header_height + 5, sum(col_widths) + 6, header_height, fill=1, stroke=0)
            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 12)
            for i, header in enumerate(headers):
                x_center = col_positions[i] + col_widths[i] / 2
                c.drawCentredString(x_center, y - 0.2*inch, header)
            y -= header_height + 2
            c.setLineWidth(1)
            c.line(margin - 3, y + 5, margin - 3 + sum(col_widths) + 6, y + 5)
            y -= 0.1 * inch
            c.setFont("Helvetica", 10)

        # Datos de la fila
        c.drawCentredString(col_positions[0] + col_widths[0]/2, y, compra.numero_factura)

        # Proveedor con ajuste automático
        x_centro_proveedor = col_positions[1] + col_widths[1] / 2
        max_prov_width = col_widths[1] - 8
        ajustar_y_centrar_texto(c, compra.proveedor.nombre, x_centro_proveedor, y, max_prov_width)

        c.drawCentredString(col_positions[2] + col_widths[2]/2, y, compra.fecha_compra.strftime('%d/%m/%Y'))
        c.drawRightString(col_positions[3] + col_widths[3] - 2, y, f"{compra.subtotal:.2f}")
        c.drawRightString(col_positions[4] + col_widths[4] - 2, y, f"{compra.iva:.2f}")
        c.drawRightString(col_positions[5] + col_widths[5] - 2, y, f"{compra.total:.2f}")
        y -= line_height

    # Línea final
    y -= 0.2 * inch
    c.setLineWidth(1)
    c.line(margin - 3, y + 5, margin - 3 + sum(col_widths) + 6, y + 5)
    y -= 0.4 * inch

    # Total IVA pagado alineado a la derecha
    c.setFont("Helvetica-Bold", 12)
    total_text = f"Total IVA pagado: ${total_iva:.2f}"
    c.drawRightString(margin - 3 + sum(col_widths), y, total_text)

    c.showPage()
    c.save()

    return response

@login_required
@cargo_requerido('admin' )
def reporte_compras_sri(request):
    # Valores por defecto al mes y año actual
    hoy = timezone.now()
    mes = request.GET.get('mes', str(hoy.month))
    anio = request.GET.get('anio', str(hoy.year))

    # Filtrar compras por mes y año
    compras = Compra.objects.filter(
        fecha_compra__year=anio,
        fecha_compra__month=mes
    ).order_by('fecha_compra')

    resumen = compras.aggregate(
        subtotal=Sum('subtotal'),
        iva=Sum('iva'),
        total=Sum('total')
    )

    # Si no hay compras, evitar None en el template
    resumen = {
        'subtotal': resumen['subtotal'] or 0,
        'iva': resumen['iva'] or 0,
        'total': resumen['total'] or 0,
    }

    context = {
        'compras': compras,
        'mes': int(mes),
        'anio': int(anio),
        'resumen': resumen,
    }
    return render(request, 'administrativo/compra/reporte_sri.html', context)